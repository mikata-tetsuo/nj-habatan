#!/usr/bin/env python3
"""
Claude API を使って英検速読トレーナー用パッセージを自動生成するスクリプト。

使い方:
  python generate_passages.py --grade 5 --count 10
  python generate_passages.py --grade all --count 5   # 各級5問ずつ
  python generate_passages.py --grade 3 --count 3 --theme 健康・医療

生成後は embed_passages.py を実行して HTML に反映してください。
"""
import argparse
import itertools
import json
import os
import re
import sys

import anthropic
import openpyxl

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"), override=True)
except ImportError:
    pass

# ------------------------------------------------------------------ 級仕様
GRADE_SPECS = {
    "5": {
        "name": "英検5級",
        "word_count": "20〜35語",
        "target_time": 15,
        "eiken_grade": "5",
        "sub_level": 1,
        "description": "小学生〜中学初級レベル。超日常的な話題（家族・食事・学校・ペット・日課）。",
        "vocab": "中学1年レベルの最基本語彙のみ。be動詞・一般動詞の現在形・過去形のみ。難しい単語は絶対に使わない。",
    },
    "4": {
        "name": "英検4級",
        "word_count": "50〜70語",
        "target_time": 30,
        "eiken_grade": "4",
        "sub_level": 1,
        "description": "中学生レベル。学校・旅行・趣味・スポーツ・食べ物・動物・地域の話題。AIや環境問題・ビジネスは使わない。",
        "vocab": "中学基本語彙のみ。専門用語・抽象概念・難語彙は一切使わない。",
    },
    "3": {
        "name": "英検3級",
        "word_count": "100〜130語",
        "target_time": 55,
        "eiken_grade": "3",
        "sub_level": 1,
        "description": "中学卒業レベル。社会問題（環境・テクノロジー・文化等）の入門的内容。",
        "vocab": "中学修了程度の語彙。接続詞や関係代名詞を使った文が多い。",
    },
    "pre2": {
        "name": "英検準2級",
        "word_count": "130〜160語",
        "target_time": 72,
        "eiken_grade": "pre2",
        "sub_level": 1,
        "description": "高校中級レベル。科学・社会・文化などの一般的なトピック。",
        "vocab": "高校中級語彙。譲歩・対比・因果の構造が頻出。",
    },
    "pre2p": {
        "name": "英検準2級+（準2〜2級の橋渡し）",
        "word_count": "160〜190語",
        "target_time": 85,
        "eiken_grade": "pre2",
        "sub_level": 2,
        "description": "準2〜2級の橋渡しレベル。社会問題・ビジネス・科学等のやや高度な内容。",
        "vocab": "高校上級語彙。複雑な文構造と抽象的な概念が登場。",
    },
    "2": {
        "name": "英検2級",
        "word_count": "180〜220語",
        "target_time": 93,
        "eiken_grade": "2",
        "sub_level": 1,
        "description": "高校卒業レベル。学術的な話題（心理・経済・環境・医療等）。",
        "vocab": "大学入試レベルの語彙。学術的な論述スタイル。",
    },
    "pre1": {
        "name": "英検準1級",
        "word_count": "220〜260語",
        "target_time": 115,
        "eiken_grade": "pre1",
        "sub_level": 1,
        "description": "大学中級レベル。専門性の高いトピック（認知科学・政策・環境経済等）。",
        "vocab": "上級語彙・学術専門用語。複雑な論理構造。",
    },
}

THEMES = [
    "環境・気候変動",
    "テクノロジー・AI",
    "健康・医療・科学",
    "教育・学習",
    "社会問題・格差",
    "経済・ビジネス",
    "文化・歴史",
    "食・農業",
    "スポーツ・芸術",
    "宇宙・自然",
    "心理・行動科学",
    "エネルギー・インフラ",
    "グローバル化・国際関係",
    "言語・コミュニケーション",
    "動物・生態系",
]

HEADERS = [
    "id", "type", "eiken_grade", "sub_level", "passage", "japanese_translation",
    "question", "choice_a", "choice_b", "choice_c", "choice_d", "answer",
    "target_time_sec", "reading_point", "evidence_text", "reward_point",
]

PROMPT = """\
英検速読トレーナー用の読解問題を1問作成してください。

## 仕様
- レベル: {name}
- 語数: {word_count}
- テーマ: {theme}
- 内容の難易度: {description}
- 語彙レベル: {vocab}

## 出力形式（JSONのみ。説明文・コードブロック不要）
{{
  "passage": "英文パッセージ",
  "japanese_translation": "日本語訳（原文に忠実に）",
  "question": "英語の問題文（What / Why / How 等で始まる）",
  "choice_a": "選択肢A",
  "choice_b": "選択肢B",
  "choice_c": "選択肢C",
  "choice_d": "選択肢D",
  "answer": "正解（a / b / c / d のいずれか1文字・小文字）",
  "reading_point": "日本語での読解ヒント（どこに注目すればよいか）",
  "evidence_text": "答えの根拠となる本文の一節（英語）"
}}

## 制約
- 不正解の選択肢は「本文にない情報」または「本文と逆の内容」にする
- reading_point は日本語で書く（英語不可）
- answer は小文字1文字（a / b / c / d）
- テーマはあくまでもヒント。自然な英文になるなら多少ずれてもよい
"""

PROMPT_5 = """\
英検5級レベルの超簡単な読解問題を1問作成してください。

## 形式（以下の2種類をランダムに選んで使うこと）
A) 3文程度の短い英文（日常場面の説明文）
B) 2人の短い会話文（A: 〜  B: 〜 形式）

## 仕様
- 語数: 20〜35語
- テーマヒント: {theme}
- テーマ例: 家族・食事・学校・ペット・日課・趣味・天気・買い物・スポーツ（超日常的なものだけ）
- 語彙: 中学1年の最基本語彙のみ（難しい単語は絶対に使わない）
- 文法: be動詞・一般動詞の現在形・過去形・疑問文のみ
- 質問: 「What / Who / Where / When does〜?」レベルの簡単な内容確認問題

## 出力形式（JSONのみ。説明文・コードブロック不要）
{{
  "passage": "英文（20〜35語）",
  "japanese_translation": "日本語訳（原文に忠実に）",
  "question": "英語の問題文（What / Who / Where 等で始まる）",
  "choice_a": "選択肢A（2〜5語程度の短い答え）",
  "choice_b": "選択肢B",
  "choice_c": "選択肢C",
  "choice_d": "選択肢D",
  "answer": "正解（a / b / c / d のいずれか1文字・小文字）",
  "reading_point": "日本語でのヒント（例：「〜という部分に注目！」）",
  "evidence_text": "答えの根拠となる本文の一節（英語）"
}}

## 制約
- 難しい単語・専門用語・抽象的な概念は絶対に使わない
- 選択肢も短くシンプルに（名詞句・短文）
- answer は小文字1文字（a / b / c / d）
"""


# ------------------------------------------------------------------ ユーティリティ
def load_excel(path: str) -> list[dict]:
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        if not row[0]:
            continue
        d = {h: (row[i] if i < len(row) else "") for i, h in enumerate(headers)}
        data.append(d)
    return data


def save_excel(path: str, all_data: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "passages"
    ws.append(HEADERS)
    for d in all_data:
        ws.append([d.get(h, "") for h in HEADERS])
    wb.save(path)


def parse_json(text: str) -> dict | None:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r'\{.*\}', text, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except json.JSONDecodeError:
                pass
    return None


# ------------------------------------------------------------------ 生成
def generate_one(client: anthropic.Anthropic, grade_key: str, theme: str,
                 next_id: int) -> dict | None:
    spec = GRADE_SPECS[grade_key]
    if grade_key == "5":
        prompt = PROMPT_5.format(theme=theme)
    else:
        prompt = PROMPT.format(theme=theme, **spec)

    try:
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
    except Exception as e:
        print(f"[API ERROR] {e}")
        return None

    data = parse_json(response.content[0].text.strip())
    if data is None:
        print("[ERROR] JSON解析失敗")
        return None

    required = ["passage", "japanese_translation", "question",
                "choice_a", "choice_b", "choice_c", "choice_d",
                "answer", "reading_point", "evidence_text"]
    missing = [f for f in required if f not in data]
    if missing:
        print(f"[ERROR] フィールド不足: {missing}")
        return None

    data.update({
        "id": next_id,
        "type": "reading",
        "eiken_grade": spec["eiken_grade"],
        "sub_level": spec["sub_level"],
        "target_time_sec": spec["target_time"],
        "reward_point": 10,
    })
    return data


# ------------------------------------------------------------------ メイン
def main():
    parser = argparse.ArgumentParser(description="英検速読トレーナー 問題自動生成")
    parser.add_argument("--grade", required=True,
                        help="生成する級: 5 / 4 / 3 / pre2 / pre2p / 2 / pre1 / all")
    parser.add_argument("--count", type=int, default=5,
                        help="1つの級に生成する問題数 (default: 5)")
    parser.add_argument("--theme", default=None,
                        help="テーマを固定したい場合に指定（省略時はローテーション）")
    parser.add_argument("--output", default="英検速読トレーナー_DB_v1.xlsx",
                        help="出力先 Excel ファイル")
    parser.add_argument("--replace", action="store_true",
                        help="指定した級の既存データを削除してから生成・保存する")
    args = parser.parse_args()

    grades = list(GRADE_SPECS.keys()) if args.grade == "all" else [args.grade]
    for g in grades:
        if g not in GRADE_SPECS:
            print(f"[ERROR] 不明な級: {g}  有効値: {list(GRADE_SPECS.keys())}")
            sys.exit(1)

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("[ERROR] ANTHROPIC_API_KEY が設定されていません")
        print("  sokutore/.env に ANTHROPIC_API_KEY=xxx を記入してください")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), args.output)
    existing = load_excel(output_path)

    if args.replace:
        before = len(existing)
        existing = [d for d in existing if str(d.get("eiken_grade","")).strip() not in grades
                    or (len(grades) > 1 and str(d.get("eiken_grade","")).strip() not in grades)]
        # gradeがpre2p(sub_level=2)の場合は eiken_grade=pre2 のsub_level=2を削除
        def should_remove(d):
            g_val = str(d.get("eiken_grade","")).strip()
            sl    = str(d.get("sub_level","")).strip()
            for gk in grades:
                spec = GRADE_SPECS[gk]
                if g_val == spec["eiken_grade"] and sl == str(spec["sub_level"]):
                    return True
            return False
        existing = [d for d in load_excel(output_path) if not should_remove(d)]
        print(f"[replace] {before - len(existing)}件削除 (残:{len(existing)}件)")

    # IDを振り直す（全データ連番）

    new_passages: list[dict] = []
    theme_iter = itertools.cycle([args.theme] if args.theme else THEMES)
    def to_int(v):
        try: return int(v)
        except: return 0
    next_id = max((to_int(d.get("id", 0)) for d in existing), default=-1) + 1
    if args.replace:
        # 既存データのIDを0から振り直す
        for i, d in enumerate(existing):
            d["id"] = i
        next_id = len(existing)

    for grade_key in grades:
        print(f"\n=== {GRADE_SPECS[grade_key]['name']} ({args.count}問) ===")
        for i in range(args.count):
            theme = next(theme_iter)
            print(f"  [{i+1:>3}/{args.count}] {theme} ... ", end="", flush=True)
            result = generate_one(client, grade_key, theme, next_id)
            if result:
                new_passages.append(result)
                next_id += 1
                print(f"OK  (id={result['id']})")
            else:
                print("スキップ")

    if not new_passages:
        print("\n生成された問題がありませんでした。")
        return

    all_data = existing + new_passages
    save_excel(output_path, all_data)
    print(f"\n[完了] {len(new_passages)}問を追記 -> {output_path}")
    print("次のコマンドで HTML に反映してください:")
    print("  python embed_passages.py")


if __name__ == "__main__":
    main()
